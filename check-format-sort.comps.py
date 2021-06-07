from openpyxl import*
from datetime import *
from datetime import timedelta

folder = "C:\\Users\\betty\\Desktop\\USING\\Comps\\"
names = ["200W Comp. 60 Days", "BPC Comp. 60 Days", "Court Sq Comp. 60 Days", "Downtown Brooklyn Comp. 60 Days", "Fulton St Comp. 60 Days", "Hudson Yards Comp. 60 Days", "LIC Waterfront Comp. 60 Days", "MWC Comp. 60 Days", "West Village Comp. 60 Days"] 

for buildingName in names:
    wb = load_workbook(filename = folder + buildingName + '.xlsx') 
    for sheet in wb.worksheets:
        if sheet.title == "Total Rental Overview" or sheet.title == "Rental Overview":
            continue

        oldDateRowsList = []
        emptyRowsList = {}
        rowNum = 1 
        for listEachRow in sheet.iter_rows(min_row=1, max_col=8):
            if (listEachRow[0].value) == "START":
                oldDataEnd = rowNum
                break
            rowNum += 1
        newDataStart = oldDataEnd + 1
        rowNum = newDataStart

        table = sheet._tables[0]
        tableDims = table.ref
        tableStartRow = int(tableDims.split(":")[0][1:]) + 1
        tableFirstRow = int(tableDims.split(":")[0][1:]) + 1
        tableLastRow = int(tableDims.split(":")[1][1:])
        newDataRows = sheet.max_row - oldDataEnd

        for eachRow in sheet.iter_rows(min_row=newDataStart, max_col=7):
            if eachRow[1].value is None: 
                break
            else:
                unitSplit = ((eachRow[1].value).split(" - "))[0]
                sheet.cell(row=eachRow[1].row, column=2, value=unitSplit)
                priceSplit = ((str(eachRow[2].value)).split("\xa0"))[0].replace("$","").replace(",","")
                sheet.cell(row=eachRow[2].row, column=3).value = float(priceSplit)
                sheet.cell(row=eachRow[2].row, column=3).number_format = '$* #,##0.00_'
                
                try:
                    bedsSplit = ((str(eachRow[3].value)).split(", "))[1]
                except:
                    bedsSplit = eachRow[3].value
                sheet.cell(row=eachRow[3].row, column=4).value = bedsSplit

                baths = sheet.cell(row=eachRow[4].row, column=5).value
                
                try:
                    SQFTSplit = ((str(eachRow[5].value)).split("\xa0"))
                    if len(SQFTSplit) == 1:
                        SQFTSplit = ((str(eachRow[5].value)).split(" "))[0]
                    else:
                        SQFTSplit = SQFTSplit[0]
                    SQFTSplit = SQFTSplit.replace(',','')
                except:
                    SQFTSplit = 0
                if SQFTSplit == 'None':
                    SQFTSplit = 0    
                sheet.cell(row=eachRow[5].row, column=6).value = float(SQFTSplit)

                if tableStartRow < (newDataRows + tableFirstRow):
                    sheet.cell(row=tableStartRow, column=1).value = eachRow[0].value
                    sheet.cell(row=tableStartRow, column=2).value = unitSplit
                    sheet.cell(row=tableStartRow, column=4).value = float(priceSplit)
                    sheet.cell(row=tableStartRow, column=4).number_format = '$* #,##0.00'
                    sheet.cell(row=tableStartRow, column=5).value = bedsSplit
                    sheet.cell(row=tableStartRow, column=6).value = baths                  
                    sheet.cell(row=tableStartRow, column=7).value = float(SQFTSplit)

            tableStartRow += 1
            rowNum += 1

            for eachRow in sheet.iter_rows(min_row=tableStartRow, max_col=7, max_row=tableLastRow):
                    sheet.cell(row=eachRow[1].row, column=1).value = None
                    sheet.cell(row=eachRow[1].row, column=2).value = None
                    sheet.cell(row=eachRow[1].row, column=4).value = None
                    sheet.cell(row=eachRow[1].row, column=5).value = None
                    sheet.cell(row=eachRow[1].row, column=6).value = None
                    sheet.cell(row=eachRow[1].row, column=7).value = None

        for eachRow in sheet.iter_rows(min_row=newDataStart, max_col=6, max_row=sheet.max_row):
                sheet.cell(row=eachRow[1].row, column=1).value = None
                sheet.cell(row=eachRow[1].row, column=2).value = None
                sheet.cell(row=eachRow[1].row, column=3).value = None
                sheet.cell(row=eachRow[1].row, column=4).value = None
                sheet.cell(row=eachRow[1].row, column=5).value = None
                sheet.cell(row=eachRow[1].row, column=6).value = None
                    
        wb.save(folder + buildingName + '2' + '.xlsx')

print(datetime.now())
